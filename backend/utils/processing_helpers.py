from quart import jsonify
from io import BytesIO
import os
import json

import pandas as pd
from utils.general import get_cur_time
from utils.enums import Templates
import asyncio
from utils.general import read_json, write_json_async, get_download_path, delete_files, write_to_json, read_from_json, get_insurance_total
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utils.templates_helpers import get_template_path_from_type

from utils.enums import Paths, Accounting

# for post, ilac and EAPC
column_map = {
        "Student ID": "Student #",
        "First Name": "First Name*",
        "Last Name": "Last Name*",
        "Birthdate": "Birthdate*",
        "Gender": "Gender*",
        "Country of Citizenship": "Country of Origin*", 
        "Legacy Email": "Insured's Primary Email*"
}

# for ININ accounting reports...
accounting_column_map = {
        "Selected Term Desc": "Selected Term Desc",
        "Student ID": "Student ID",
        "First Name": "First Name",
        "Last Name": "Last Name",
        "Balance": "Balance"
        # balance column -> +20 or -20
        # Fee Code column -> ININ
}



#------------------------------------------------------------




async def populate_ESL(esl_eapc: pd.DataFrame):

    # ESL EAPC - 'major' column has value of 'ESL EAPC'
    
        #----
        # Grab ESL EAPC -> template pointer
        # get DF
        #esl_eapc_template_df = pd.read_excel("./data/templates/insurance/ESL EAP GuardMe template.xlsx")

        # Create a NEW empty template-shaped dataframe
        #output_df = pd.DataFrame(columns=esl_eapc_template_df.columns)

    

    try:

        print("length of DF given to ESL EAPC before: ", len(esl_eapc))
        # template_name = await find_template_by_keyword(Templates.ESL.value, Paths.INSURANCE_TEMPLATES.value)
        # if template_name == None:
        #     return jsonify({"error": "Populate ESL keyword for template no match"}), 500



        if os.path.exists("/tmp/data/config.json"):
            with open("/tmp/data/config.json", "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                    print("Config.json contents:", json.dumps(data, indent=4), flush=True)
                except json.JSONDecodeError as e:
                    print("Config.json exists but is invalid JSON!", flush=True)
                    raw = f.read()
                    print("Raw contents:", repr(raw), flush=True)
        else:
            print("Config.json file does not exist!", flush=True)

        
        
        template_path, key = get_template_path_from_type(Templates.ESL.value)
        template_name = await read_from_json(Paths.TEMPLATES_KEY.value, key)
        final_path = template_path + template_name

        print("right before load_workbook after final_path var is created.")

        wb = load_workbook(final_path)
        ws = wb.active

        # sanity checks
        if esl_eapc.empty:
            print("DEBUG: source dataframe (esl_eapc) is empty — nothing to write.")
        else:
            print(f"DEBUG: source dataframe has {len(esl_eapc)} rows")

        # eventually grabbed from json...
        HEADER_ROW = 13

        # Build a mapping of template header name -> column letter
        # {Student # : C}
        template_columns = {}
        for col_idx, cell in enumerate(ws[HEADER_ROW], start=1):  # header row is ws[1]
            if cell.value:  # skip empty or merged secondary cells
                template_columns[cell.value] = get_column_letter(col_idx)


        # Populate rows from source into template
        current_row = HEADER_ROW + 1  # 14
        #print(list(template_columns.items()))

        for _, df_row in esl_eapc.iterrows():
            for source_col, template_col_name in column_map.items():
                col_letter = template_columns.get(template_col_name)
                country_col_letter = template_columns.get("Country*")
                city_col_letter = template_columns.get("City*")
                #print(city_col_letter)
                if col_letter and source_col in df_row.index:
                    #print(f"writing '{df_row[source_col]}' to col {col_letter} row {current_row}")
                    ws[f"{col_letter}{current_row}"] = df_row[source_col]
                    # write into Country
                    ws[f"{country_col_letter}{current_row}"] = "Canada"
                    # write into City
                    ws[f"{city_col_letter}{current_row}"] = "Barrie"
            current_row += 1
        
        # auto-populate - City = Barrie
        # auto-populate - Country = Canada

        folder_path = Templates.POP_TEMPLATE_PATH.value + Templates.ESL.value + "/"

        # destroy other templates that exist in /ESL first
        delete_files(folder_path)

        # write new populated ESL EAPC .xlsx file to directory
        cur_time = get_cur_time()
        new_esl_report_name = cur_time + "_ESL_EAPC_insurance_report.xlsx"
        file_path = folder_path + new_esl_report_name
        
        #save new ouput
        wb.save(file_path)

        # write to config.json
        await write_to_json(new_esl_report_name, Templates.TEMPLATE_CONFIG_KEY.value, Templates.ESL.value)

        # Write new ESL populated template to config.json
        # settings = await asyncio.to_thread(read_json, Paths.CONFIG_PATH.value)
        # print("after we read config json settings full_process")
        # print(Templates.TEMPLATE_CONFIG_KEY.value)
        # # change json line to new name
        # settings[Templates.TEMPLATE_CONFIG_KEY.value][Templates.ESL.value] = new_esl_report_name
        # # run coroutine task, to write to json for new baseline data
        # asyncio.create_task(write_json_async(Paths.CONFIG_PATH.value, settings))
        
        # print("ESL populated successfully")
        return True
    
    #return exception
    except Exception as error:
        print(error)
        return False
    

# ILAC populate
async def populate_ILAC(df: pd.DataFrame):

    # ILAC - 'Campus' column has value of 'LT'
        # auto-populate - City = Toronto
        # auto-populate - Country = Canada
        # ILAC GuardMe template.xlsx
    
    try:
        # print("running populate ILAC")
        print("length of DF given to ILAC before: ", len(df))

        # read config.json if it exists ...
        if os.path.exists("/tmp/data/config.json"):
            with open("/tmp/data/config.json", "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                    print("Config.json contents:", json.dumps(data, indent=4), flush=True)
                except json.JSONDecodeError as e:
                    print("Config.json exists but is invalid JSON!", flush=True)
                    raw = f.read()
                    print("Raw contents:", repr(raw), flush=True)
        else:
            print("Config.json file does not exist!", flush=True)

        template_path, key = get_template_path_from_type(Templates.ILAC.value)
        print("key for ilac: ", key)
        template_name = await read_from_json(Paths.TEMPLATES_KEY.value, key)
        final_path = template_path + template_name

        print("right before load_workbook after final_path var is created.")

        wb = load_workbook(final_path)
        ws = wb.active

        # sanity checks
        if df.empty:
            print("DEBUG: source dataframe (ilac) is empty — nothing to write.")
        else:
            print(f"DEBUG: source dataframe has {len(df)} rows")

        # eventually grabbed from json...
        HEADER_ROW = 13

        # Build a mapping of template header name -> column letter
        # {Student # : C}
        template_columns = {}
        for col_idx, cell in enumerate(ws[HEADER_ROW], start=1):  # header row is ws[1]
            if cell.value:  # skip empty or merged secondary cells
                template_columns[cell.value] = get_column_letter(col_idx)

        # Populate rows from source into template
        current_row = HEADER_ROW + 1  # 14
        #print(list(template_columns.items()))

        for _, df_row in df.iterrows():
            for source_col, template_col_name in column_map.items():
                col_letter = template_columns.get(template_col_name)
                country_col_letter = template_columns.get("Country*")
                city_col_letter = template_columns.get("City*")
                #print(city_col_letter)
                if col_letter and source_col in df_row.index:
                    #print(f"writing '{df_row[source_col]}' to col {col_letter} row {current_row}")
                    ws[f"{col_letter}{current_row}"] = df_row[source_col]
                    # write into Country
                    ws[f"{country_col_letter}{current_row}"] = "Canada"
                    # write into City
                    ws[f"{city_col_letter}{current_row}"] = "Toronto"
            current_row += 1

        folder_path = Templates.POP_TEMPLATE_PATH.value + Templates.ILAC.value + "/"

        # destroy other templates that exist in /ESL first
        delete_files(folder_path)

        # write new populated ESL EAPC .xlsx file to directory
        cur_time = get_cur_time()
        new_ilac_report_name = cur_time + "_ILAC_GuardMe_template.xlsx"
        file_path = folder_path + new_ilac_report_name
        
        #save new ouput
        wb.save(file_path)

        # write to config.json
        await write_to_json(new_ilac_report_name, Templates.TEMPLATE_CONFIG_KEY.value, Templates.ILAC.value)

        print("ILAC populated successfully")
        return True
    
    #return exception
    except Exception as error:
        print("Error in populate_ILAC function in processing_helpers.py")
        print(error)
        return False


# POST populate
async def populate_POST(df: pd.DataFrame):

    # POST SECONDARY - All students left after previous processing filters are applied
        # auto-populate - City = Barrie
        # auto-populate - Country = Canada
        # POST SECONDARY GuardMe template.xlsx
    
    try:
        # print("running populate POST")
        print("length of DF given to post secondary before: ", len(df))

        #-----------------------------------------
        # read config.json if it exists ...
        if os.path.exists("/tmp/data/config.json"):
            with open("/tmp/data/config.json", "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                    print("Config.json contents:", json.dumps(data, indent=4), flush=True)
                except json.JSONDecodeError as e:
                    print("Config.json exists but is invalid JSON!", flush=True)
                    raw = f.read()
                    print("Raw contents:", repr(raw), flush=True)
        else:
            print("Config.json file does not exist!", flush=True)
        #-----------------------------------------




        template_path, key = get_template_path_from_type(Templates.POST.value)
        template_name = await read_from_json(Paths.TEMPLATES_KEY.value, key)
        final_path = template_path + template_name

        print("right before load_workbook after final_path var is created.")

        wb = load_workbook(final_path)
        ws = wb.active

        # sanity checks
        if df.empty:
            print("DEBUG: source dataframe (post) is empty — nothing to write.")
        else:
            print(f"DEBUG: source dataframe has {len(df)} rows")

        # eventually grabbed from json...
        HEADER_ROW = 13

        # Build a mapping of template header name -> column letter
        # {Student # : C}
        template_columns = {}
        for col_idx, cell in enumerate(ws[HEADER_ROW], start=1):  # header row is ws[1]
            if cell.value:  # skip empty or merged secondary cells
                template_columns[cell.value] = get_column_letter(col_idx)

        # Populate rows from source into template
        current_row = HEADER_ROW + 1  # 14
        #print(list(template_columns.items()))

        for _, df_row in df.iterrows():
            for source_col, template_col_name in column_map.items():
                col_letter = template_columns.get(template_col_name)
                country_col_letter = template_columns.get("Country*")
                city_col_letter = template_columns.get("City*")
                #print(city_col_letter)
                if col_letter and source_col in df_row.index:
                    #print(f"writing '{df_row[source_col]}' to col {col_letter} row {current_row}")
                    ws[f"{col_letter}{current_row}"] = df_row[source_col]
                    # write into Country
                    ws[f"{country_col_letter}{current_row}"] = "Canada"
                    # write into City
                    ws[f"{city_col_letter}{current_row}"] = "Barrie"
            current_row += 1

        folder_path = Templates.POP_TEMPLATE_PATH.value + Templates.POST.value + "/"

        # destroy other templates that exist in /ESL first
        delete_files(folder_path)

        # write new populated ESL EAPC .xlsx file to directory
        cur_time = get_cur_time()
        new_post_report_name = cur_time + "_POST_GuardMe_template.xlsx"
        file_path = folder_path + new_post_report_name
        
        #save new ouput
        wb.save(file_path)

        # write to config.json
        await write_to_json(new_post_report_name, Templates.TEMPLATE_CONFIG_KEY.value, Templates.POST.value)

        print("POST populated successfully")
        return True
    
    #return exception
    except Exception as error:
        print(error)
        return False



# Accounting populate somewhere

async def populate_accounting(df: pd.DataFrame):

    try:
        print("running populate ACCOUNTING")

        # - fee targets
            # - POST / ILAC
            # - EAPC / ESLG

        template_path, key = get_template_path_from_type(Templates.ACCOUNTING.value)
        template_name = await read_from_json(Paths.TEMPLATES_KEY.value, key)
        final_path = template_path + template_name

        print("right before load_workbook after final_path var is created.")

        # eventually grabbed from json...
        HEADER_ROW = 1

        wb = load_workbook(final_path)
        ws = wb.active

        # Get current max column
        max_col = ws.max_column
        # Add "Balance" as a new column header at the end
        ws.cell(row=HEADER_ROW, column=max_col + 1, value="Balance")

        # sanity checks
        if df.empty:
            print("DEBUG: source dataframe (ACCOUNTING) is empty — nothing to write.")
        else:
            print(f"DEBUG: source dataframe has {len(df)} rows")

        

        template_columns = {}
        for col_idx, cell in enumerate(ws[HEADER_ROW], start=1):  # header row is ws[1]
            if cell.value:  # skip empty or merged secondary cells
                template_columns[cell.value] = get_column_letter(col_idx)
        
        current_row = HEADER_ROW + 1  # 2
        print(list(template_columns.items()))

        for _, df_row in df.iterrows():
            for source_col, template_col_name in accounting_column_map.items():
                col_letter = template_columns.get(template_col_name)
                notes_letter = template_columns.get("Notes")
                #print(city_col_letter)
                if col_letter and source_col in df_row.index:
                    #print(f"writing '{df_row[source_col]}' to col {col_letter} row {current_row}")
                    ws[f"{col_letter}{current_row}"] = df_row[source_col]
                    ws[f"{notes_letter}{current_row}"] = "ININ"
            current_row += 1

        # ws[f"{city_col_letter}{current_row}"] = "Barrie"

        folder_path = Templates.POP_TEMPLATE_PATH.value + Templates.ACCOUNTING.value + "/"

        # destroy other templates that exist in /ESL first
        delete_files(folder_path)

        # write new populated ESL EAPC .xlsx file to directory
        cur_time = get_cur_time()
        new_accounting_report_name = cur_time + "_ACCOUNTING_template.xlsx"
        file_path = folder_path + new_accounting_report_name
        
        #save new ouput
        wb.save(file_path)

        # write to config.json
        await write_to_json(new_accounting_report_name, Templates.TEMPLATE_CONFIG_KEY.value, Templates.ACCOUNTING.value)

        print("ACCOUNTING populated successfully")

        return True

    #return exception
    except Exception as error:
        print(error)
        return False






def process_fees(total: float, year: str, semester: str, df: pd.DataFrame) -> pd.DataFrame:

    # year = str(year)
    #next_year = str(int(year) + 1)
    previous_year = str(int(year) - 1)

    print("starting process_fees")

    balance_rules = {
        "FALL": lambda r: (total- pd.to_numeric(r[f"Fall {year} Fees Paid"], errors="coerce")),
        "WINTER": lambda r: (
            total
            -
            (
                pd.to_numeric(r[f"Fall {previous_year} Fees Paid"], errors="coerce") +
                pd.to_numeric(r[f"Winter {year} Fees Paid"], errors="coerce")
            )
        ),

        "SUMMER": lambda r: (
            total
            -
            (
                pd.to_numeric(r[f"Fall {previous_year} Fees Paid"], errors="coerce") +
                pd.to_numeric(r[f"Winter {year} Fees Paid"], errors="coerce") +
                pd.to_numeric(r[f"Summer {year} Fees Paid"], errors="coerce")
            )
        )
    }

    if semester.upper() not in balance_rules:
        raise ValueError(f"Unknown semester: {semester}")
    
    df_copy = df.copy()
    df_copy["Balance"] = pd.Series(dtype=float)  # create balance column before we process

    balance_fn = balance_rules[semester.upper()]
    df_copy["Balance"] = df_copy.apply(balance_fn, axis=1)

    # filter through NEW balance column, if not equal to zero, then we keep for new df to return 
    unbalanced_df = df_copy[df_copy["Balance"] != 0].copy()
    print(f"len of unbalanced_df in process_fees: {len(unbalanced_df)}")
    print(unbalanced_df["Balance"])



    return unbalanced_df


async def get_fees_total(type:str, semester: str) -> float:
    print("started get fee total calc...")

    # switch for types
    total_compare = await get_insurance_total(type, semester)
    print(f"Total: {total_compare:.2f}")

    return total_compare




# get post dataframe for its target
# Return:
    # COGNOS with only fields that are +/- Fees Paid balance from POST / ILAC calculations
async def get_balance_df(df: pd.DataFrame, semester: str, year: str, type: str):
    lowered_semester = semester.lower()
    
    # total to compare against
    total = await get_fees_total(type,lowered_semester)

    result = process_fees(total, year, semester, df)
    return result


# async def get_eapc_df(df: pd.DataFrame, semester: str, year: str):

#     # year = str(year)
#     next_year = str(int(year) + 1)
#     # total to compare against
#     total = await get_fees_total("normal",semester)

#     'Winter {year} Fees Paid'
#     result = process_fees(total, year, next_year, semester, df)
#     return result



# accounting fee calculations
# Take a df ; either baseline or compare report
# Parameters:
    # df
    # semester
    # year
# RETURNS - accounting df
async def accounting_calculations(post_ilac_df: pd.DataFrame, eapc_df: pd.DataFrame, semester: str, year: str) -> pd.DataFrame:

    # logic such as which semester column to grab from;
        # Fall 2025, or winter 2026, 2025 total fees paid
    
    # POST / ILAC cognos + column 'Balance'
    post_df = await get_balance_df(post_ilac_df, semester,year, "post")
    print("'after post df first balance...")
    # EAPC cognos + column 'Balance'
    eapc_df = await get_balance_df(eapc_df, semester, year, "normal")

    # merge post + EAPC - accounting templates into one...
    # everything in here has a OWED / OWING balance
    combined_df = pd.concat([post_df, eapc_df], ignore_index=True)
    print(f"length of combined: len(combined_df)")
    
    return combined_df

